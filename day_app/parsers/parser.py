import requests
import urllib.request
import os
from urllib.parse import quote
from bs4 import BeautifulSoup
import openpyxl
from day_app.parsers.schedule import ScheduleDay, Subject, week_days, GroupSchedule
import re
import pickle


xlsx_path = 'data/xlsx'
xlsx_parsed_path = 'data/xlsx_parsed'
updates_data_file = 'data/updates.json'

regex_subject_name = re.compile(r'(?:\к\р\. (.+) \н\.)?(?:(^.+) \н\. )?(.+)')


def download_xlsx():
    page = requests.get('https://www.mirea.ru/schedule/')
    soup = BeautifulSoup(page.text, 'html.parser')

    schedule_links = [
        x['href'] for x in
        soup.find('div', {'class': 'rasspisanie'}).find(string='Институт информационных технологий'). \
        find_parent('div'). \
        find_parent('div'). \
        findAll('a', {'class': 'uk-link-toggle'})
    ]

    for i in range(len(schedule_links)):
        urllib.request.urlretrieve('https://' + quote(schedule_links[i][8:]),
                                   os.path.join(xlsx_path, f'{i + 1} курс.xlsx'))


def parse_subject_weeks_from_str(raw_value: str) -> list[int]:
    try:
        weeks: list[int] = []
        raw_value = raw_value.replace(' ', '')
        if raw_value.count(',') != 0:
            splitted_weeks = raw_value.split(',')
            for week in splitted_weeks:
                if week.count('-') != 0:
                    splitted_weeks = week.split('-')
                    weeks.extend(list(range(int(splitted_weeks[0]), int(splitted_weeks[1]) + 1)))
                else:
                    weeks.append(int(week))
        elif raw_value.count('-') != 0:
            splitted_weeks = raw_value.split('-')
            weeks.extend(list(range(int(splitted_weeks[0]), int(splitted_weeks[1]) + 1)))
        else:
            weeks.append(int(raw_value))
        return weeks
    except IndexError:
        print(raw_value)
        raise IndexError()
    except ValueError:
        print(raw_value)
        raise ValueError()


def parse_xlsx():
    xlsx_files = [os.path.join(xlsx_path, x) for x in os.listdir(xlsx_path) if x.endswith('xlsx')]
    groups_schedules: list[GroupSchedule] = []
    for xlsx in xlsx_files:
        print(xlsx)
        book = openpyxl.load_workbook(xlsx)
        sheet = book.active

        num_cols = sheet.max_column
        group_index = 0
        group_column = 0
        while group_column < num_cols:
            group_column = 6 + 5 * group_index
            group_name = sheet.cell(row=2, column=group_column).value
            if group_name is None or group_name == 'День недели':
                group_index += 1
                continue
            group_schedule = GroupSchedule(
                name=sheet.cell(row=2, column=group_column).value, schedule_days=[]
            )
            for day_index in range(6):
                schedule_day = ScheduleDay(name=week_days[day_index], periods_subjects={})
                for subject_index in range(6):
                    subject_row = 4 + day_index * 12 + subject_index * 2
                    subjects_list: list[Subject] = []
                    for i in range(2):
                        subject_cell_value = sheet.cell(row=subject_row + i, column=group_column).value
                        if subject_cell_value is None:
                            subjects_list.append(None)
                            continue
                        parsed_subject_cell = regex_subject_name.match(subject_cell_value)
                        subject = Subject(
                            test_weeks=parse_subject_weeks_from_str(parsed_subject_cell.groups()[0]) if
                            parsed_subject_cell.groups()[0] is not None else None,
                            on_weeks=parse_subject_weeks_from_str(parsed_subject_cell.groups()[1]) if
                            parsed_subject_cell.groups()[1] is not None else None,
                            name=parsed_subject_cell.groups()[2],
                            form=sheet.cell(row=subject_row + i, column=group_column + 1).value,
                            teacher=sheet.cell(row=subject_row + i, column=group_column + 2).value,
                            audience_name=sheet.cell(row=subject_row + i, column=group_column + 3).value,
                        )
                        subjects_list.append(subject)
                    schedule_day.periods_subjects[subject_index] = tuple(subjects_list)
                group_schedule.schedule_days.append(schedule_day)
            group_index += 1
            groups_schedules.append(group_schedule)
    return groups_schedules


def parse_schedule():
    #download_xlsx()
    with open('groups_data', 'wb') as f:
        pickle.dump(parse_xlsx(), f)